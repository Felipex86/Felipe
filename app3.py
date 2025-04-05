from flask import Flask, render_template, request, redirect, url_for, flash, session
import datetime
import io
import base64
import os
import requests
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'supersecretkey'
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Función para obtener el pronóstico del clima
def get_weather_forecast():
    url = "https://api.open-meteo.com/v1/forecast"
    params = {
        "latitude": -41.4975,
        "longitude": -72.3079,
        "daily": ["precipitation_sum"],
        "timezone": "America/Santiago",
        "forecast_days": 7
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        data = response.json()
        dates = data["daily"]["time"]
        precipitation = data["daily"]["precipitation_sum"]
        forecast = [
            {"date": datetime.datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m"), "precip": p}
            for d, p in zip(dates, precipitation)
        ]
        return forecast
    except Exception as e:
        print(f"Error en la API del clima: {e}")
        return None

# Función para procesar archivo PROGRAMA
def process_programa_file(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb["PROGRAMA"]
        
        # Buscar "PMontt220"
        target_row = None
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                if cell.value == "PMontt220":
                    target_row = cell.row
                    break
            if target_row:
                break
        
        if not target_row:
            return None
        
        data = []
        for col in range(5, 29):  # Columnas E a AB
            cell_value = sheet.cell(row=target_row, column=col).value
            try:
                data.append(round(float(cell_value), 2) if cell_value else 0.0)
            except:
                data.append(0.0)
        return data
    except Exception as e:
        print(f"Error procesando archivo PROGRAMA: {e}")
        return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        session.clear()
        
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('Archivo principal no seleccionado')
            return redirect(request.url)
        if not allowed_file(file.filename):
            flash('Solo se permiten archivos .xlsx')
            return redirect(request.url)

        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            wb = load_workbook(filepath, data_only=True)
        except Exception as e:
            flash(f'Error: {str(e)}')
            return redirect(request.url)

        programa_data = None
        programa_file = request.files.get('programa_file')
        if programa_file and programa_file.filename != '' and allowed_file(programa_file.filename):
            try:
                programa_filename = secure_filename(programa_file.filename)
                programa_path = os.path.join(app.config['UPLOAD_FOLDER'], programa_filename)
                programa_file.save(programa_path)
                programa_data = process_programa_file(programa_path)
            except Exception as e:
                flash(f'Error en PROGRAMA: {str(e)}')

        manual_data = []
        valid = True
        for i in range(1, 5):
            tipo = request.form.get(f'tipo_{i}', '').upper()
            hora = request.form.get(f'hora_{i}', '')
            if not re.match(r'^[PB]$', tipo) or not re.match(r'^\d{2}:\d{2}$', hora):
                valid = False
                break
            manual_data.append({'tipo': tipo, 'hora': hora})
        
        if not valid:
            flash('Datos manuales inválidos')
            return redirect(request.url)

        session['programa_data'] = programa_data
        session['manual_data'] = manual_data
        session['filename'] = filename

        return render_template('select_sheet.html', sheets=wb.sheetnames, filename=filename)
    
    return render_template('upload.html')

@app.route('/process_sheet', methods=['POST'])
def process_sheet():
    filename = session.get('filename')
    sheet_name = request.form.get('sheet')

    if not filename or not sheet_name:
        flash('Error en los datos')
        return redirect(url_for('upload_file'))

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        flash(f'Error al abrir archivo: {str(e)}')
        return redirect(url_for('upload_file'))

    if sheet_name not in wb.sheetnames:
        flash('Hoja no encontrada')
        return redirect(url_for('upload_file'))

    sheet = wb[sheet_name]

    # Buscar última fila válida
    ultima_fila = None
    for row_num in range(136, 105, -1):
        b, e, f, g = [sheet.cell(row=row_num, column=col).value for col in [2, 5, 6, 7]]
        if all(v is not None for v in [b, e, f, g]):
            ultima_fila = {'B': b, 'E': e, 'F': f, 'G': g}
            break

    if not ultima_fila:
        flash('No se encontraron datos completos')
        return redirect(url_for('upload_file'))

    # Generar gráfico
    b_values = [sheet.cell(row=row_num, column=2).value for row_num in range(106, 137)]
    b_values = [v for v in b_values if isinstance(v, (int, float))]
    while len(b_values) < 7:
        b_values.insert(0, 0)
    b_values = b_values[-7:]

    plt.figure()
    plt.plot(range(1, 8), b_values, marker='o')
    plt.xlabel('Días')
    plt.ylabel('Valores')
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    plot_url = base64.b64encode(buf.getvalue()).decode('utf-8')
    plt.close()

     # ==========================
    # Extraer datos de hidrología (últimos 7 con datos)
    # ==========================
    hidro_datos_validos = []

    for row in range(407, 438):  # A407:B437
        fecha = sheet.cell(row=row, column=1).value
        agua = sheet.cell(row=row, column=2).value

        if isinstance(agua, (int, float)):
            fecha_str = fecha.strftime("%d/%m") if isinstance(fecha, datetime.datetime) else str(fecha)
            hidro_datos_validos.append({"fecha": fecha_str, "agua": round(agua, 2)})

    # Tomar las últimas 7 con datos, rellenar con ceros al principio si son menos
    hidro_ultimos_7 = hidro_datos_validos[-7:]
    while len(hidro_ultimos_7) < 7:
        hidro_ultimos_7.insert(0, {"fecha": "-", "agua": 0.0})

    # Total mensual (B438)
    try:
        total_mes = sheet.cell(row=438, column=2).value
        total_mes = round(float(total_mes), 2) if total_mes is not None else 0.0
    except:
        total_mes = 0.0

    return render_template('result.html',
                           ultima_fila=ultima_fila,
                           ultima_semana=enumerate(b_values, 1),
                           plot_url=plot_url,
                           forecast=get_weather_forecast(),
                           programa_data=session.get('programa_data'),
                           manual_data=session.get('manual_data'),
                           hidrologia=hidro_ultimos_7,
                           total_hidrologia=total_mes)

if __name__ == '__main__':
    app.run(debug=True)
